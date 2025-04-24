import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random
import logging
import re
from pathlib import Path

# 设置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# 设置用户代理，模拟浏览器访问
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0'
]

def clean_product_number(product_number):
    """清理和标准化商品货号"""
    if not product_number:
        return None
    
    # 转换为字符串
    product_number = str(product_number)
    
    # 移除空格
    product_number = product_number.strip()
    
    # 获取纯数字版本（移除小数点等）
    clean_number = ''.join(filter(str.isdigit, product_number))
    
    logging.info(f"清理货号: 原始值 -> {product_number}, 纯数字 -> {clean_number}")
    return clean_number

def get_product_details(product_number):
    """获取商品详细信息，包括原价和促销价"""
    try:
        # 保存原始货号格式用于搜索
        original_format = str(product_number).strip()
        
        # 获取清理后的纯数字货号
        clean_number = clean_product_number(product_number)
        if not clean_number:
            logging.error(f"无效的货号: {product_number}")
            return {
                "product_number": product_number,
                "original_price": None,
                "current_price": None,
                "is_on_sale": False
            }
        
        # 尝试多种可能的URL
        urls = [
            f"https://www.ikea.cn/cn/zh/search/products/?q={original_format}&qtype=search_keywords",  # 使用原始格式(带点)
            f"https://www.ikea.cn/cn/zh/search/products/?q={clean_number}&qtype=search_keywords",   # 使用纯数字格式
            f"https://www.ikea.cn/cn/zh/p/-{clean_number}/"  # 通用URL模式
        ]
        
        headers = {
            'User-Agent': random.choice(USER_AGENTS),
            'Accept': 'text/html,application/xhtml+xml,application/xml',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8'
        }
        
        response = None
        successful_url = None
        
        for url in urls:
            try:
                logging.info(f"尝试URL: {url}")
                response = requests.get(url, headers=headers, timeout=10)
                if response.status_code == 200:
                    successful_url = url
                    logging.info(f"成功获取页面: {url}")
                    break
            except Exception as e:
                logging.warning(f"URL {url} 访问失败: {str(e)}")
        
        if not response or response.status_code != 200:
            logging.error(f"无法获取商品 {product_number} 页面")
            return {
                "product_number": product_number,
                "original_price": None,
                "current_price": None,
                "is_on_sale": False
            }
        
        soup = BeautifulSoup(response.text, 'html.parser')
        html_text = response.text
        
        # 修改价格提取逻辑 - 使用更精确的正则表达式
        # 查找形如 ¥1499.00 或 ¥1,499.00 的价格
        price_pattern = r'¥\s*([\d,]+(?:\.\d{2})?)'
        matches = re.findall(price_pattern, html_text)
        
        if matches:
            # 移除逗号并转换为浮点数
            prices = []
            for match in matches:
                try:
                    # 移除逗号
                    clean_price = match.replace(',', '')
                    prices.append(float(clean_price))
                except ValueError:
                    continue
            
            # 过滤有效价格并排序
            valid_prices = [p for p in prices if p > 0]
            unique_prices = sorted(set(valid_prices))
            
            logging.info(f"页面中找到的所有价格: {unique_prices}")
            
            # 初始化价格变量
            original_price = None
            current_price = None
            
            # 如果有多个价格，尝试识别原价和现价
            if len(unique_prices) >= 2:
                # 检查是否有一个价格是下面情况之一：
                # 1. 非常低（如1元、2元），这可能是错误
                # 2. 与其他价格差异非常大
                
                # 如果最低价格小于10元且与第二低价格差异很大，可能是错误
                if unique_prices[0] < 10 and unique_prices[1] / unique_prices[0] > 10:
                    logging.warning(f"检测到异常低价: {unique_prices[0]}，忽略此价格")
                    # 使用第二低和最高价格
                    current_price = unique_prices[1]
                    original_price = unique_prices[-1]
                else:
                    # 正常情况：最低价是现价，最高价是原价
                    current_price = unique_prices[0]
                    original_price = unique_prices[-1]
                
                logging.info(f"选择 - 原价: {original_price}, 现价: {current_price}")
            elif len(unique_prices) == 1:
                # 只有一个价格时，原价和现价相同
                original_price = unique_prices[0]
                current_price = unique_prices[0]
                logging.info(f"只找到一个价格: {current_price}")
        else:
            logging.warning(f"没有找到任何价格信息")
            return {
                "product_number": product_number,
                "original_price": None,
                "current_price": None,
                "is_on_sale": False
            }
        
        # 判断是否促销
        is_on_sale = False
        if original_price and current_price and original_price > current_price:
            # 不要标记差异太大的价格为促销（可能是数据错误）
            if current_price < 10 and original_price / current_price > 100:
                logging.warning(f"价格差异过大，可能是错误数据: 原价 {original_price}, 现价 {current_price}")
                is_on_sale = False
                # 将现价设置为与原价相同，避免错误数据
                current_price = original_price
            else:
                is_on_sale = True
        
        # 从页面标签确认是否促销
        sale_indicators = ["优惠有效期", "更低价格", "会员价", "限时", "促销", "特价"]
        for indicator in sale_indicators:
            if indicator in html_text:
                logging.info(f"找到促销指标: {indicator}")
                if original_price and current_price and original_price > current_price:
                    # 再次检查价格差异是否合理
                    if current_price < 10 and original_price / current_price > 100:
                        logging.warning(f"尽管有促销指标，但价格差异过大: 原价 {original_price}, 现价 {current_price}")
                        is_on_sale = False
                        current_price = original_price
                    else:
                        is_on_sale = True
                    break
        
        return {
            "product_number": product_number,
            "original_price": original_price,
            "current_price": current_price,
            "is_on_sale": is_on_sale,
            "url": successful_url
        }
            
    except Exception as e:
        logging.error(f"获取商品 {product_number} 详细信息时出错: {str(e)}")
        return {
            "product_number": product_number,
            "original_price": None,
            "current_price": None,
            "is_on_sale": False
        }

def update_excel_prices(excel_file):
    """从Excel读取商品货号，获取当前价格并填入到现价列"""
    try:
        # 加载Excel文件
        df = pd.read_excel(excel_file)
        
        # 确保必要的列存在
        required_columns = ['订单号', '商品货号', '数量', '商品单价', '现价']
        for col in required_columns:
            if col not in df.columns:
                logging.error(f"Excel表格中缺少必要的列: {col}")
                return False
        
        # 创建工作簿对象
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 查找列的索引
        header_row = 1  # 表头行号
        product_code_col = None
        current_price_col = None
        unit_price_col = None
        
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value == '商品货号':
                product_code_col = col
            elif cell_value == '现价':
                current_price_col = col
            elif cell_value == '商品单价':
                unit_price_col = col
        
        if not product_code_col or not current_price_col or not unit_price_col:
            logging.error(f"无法找到'商品货号'、'现价'或'商品单价'列")
            return False
            
        logging.info(f"列索引: 商品货号={product_code_col}, 现价={current_price_col}, 商品单价={unit_price_col}")
        
        # 黄色填充样式（用于标记价格变化）
        yellow_fill = PatternFill(start_color='FFFF00',
                                 end_color='FFFF00',
                                 fill_type='solid')
        
        # 处理每个商品
        updated_count = 0
        price_change_count = 0
        
        # 从第2行开始（跳过表头）
        for row in range(2, ws.max_row + 1):
            product_code = ws.cell(row=row, column=product_code_col).value
            
            # 跳过空行和自提/物流货号
            if not product_code or str(product_code).startswith('500.'):
                continue
                
            product_code = str(product_code).strip()
            logging.info(f"处理第 {row} 行，商品货号: {product_code}")
            
            # 获取商品当前价格
            details = get_product_details(product_code)
            current_price = details.get('current_price')
            
            # 如果成功获取价格，更新Excel
            if current_price:
                # 获取现有价格（如果有）
                existing_price_cell = ws.cell(row=row, column=current_price_col)
                existing_price = existing_price_cell.value
                
                # 获取商品单价
                unit_price_cell = ws.cell(row=row, column=unit_price_col)
                unit_price = unit_price_cell.value
                
                # 更新价格
                existing_price_cell.value = current_price
                updated_count += 1
                
                # 检查价格是否变化
                if existing_price and existing_price != current_price:
                    logging.info(f"价格变化: {product_code} - 原来: {existing_price}, 现在: {current_price}")
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                    price_change_count += 1
                
                # 【新功能】检查现价是否低于商品单价，如果是则标记为黄色
                if unit_price and current_price < unit_price:
                    logging.info(f"现价低于商品单价: {product_code} - 单价: {unit_price}, 现价: {current_price}")
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = yellow_fill
                    
                logging.info(f"更新商品 {product_code} 的现价为 {current_price}")
            else:
                logging.warning(f"无法获取商品 {product_code} 的价格")
                
            # 随机延迟，避免被网站封锁
            time.sleep(random.uniform(1, 3))
        
        # 保存更新后的Excel文件
        wb.save(excel_file)
        logging.info(f"Excel更新完成。共更新 {updated_count} 个价格，{price_change_count} 个价格有变化。")
        
        return True
    except Exception as e:
        logging.error(f"更新Excel时出错: {str(e)}")
        return False

def test_single_product(product_number):
    """测试单个商品的价格获取"""
    print(f"\n正在测试商品 {product_number}...")
    details = get_product_details(product_number)
    print(f"商品: {product_number}")
    print(f"原价: {details.get('original_price')}")
    print(f"现价: {details.get('current_price')}")
    print(f"是否促销: {details.get('is_on_sale')}")
    if 'url' in details:
        print(f"成功URL: {details.get('url')}")
    return details

def test_with_unit_price(product_number, unit_price):
    """测试单个商品的价格获取并与单价比较"""
    print(f"\n正在测试商品 {product_number}，单价 {unit_price}...")
    details = get_product_details(product_number)
    current_price = details.get('current_price')
    print(f"商品: {product_number}")
    print(f"单价: {unit_price}")
    print(f"现价: {current_price}")
    
    if current_price and current_price < unit_price:
        print(f"【价格下降】现价低于单价: 差额 {unit_price - current_price:.2f}")
    elif current_price:
        print(f"现价未低于单价")
    else:
        print(f"无法获取现价")
    
    return details

if __name__ == "__main__":
    # Excel文件路径
    excel_file = "F:\\宜家自动查询\\订单汇总.xlsx"  # 订单信息Excel文件路径
    
    # 是否进行测试模式
    TEST_MODE = False
    
    if TEST_MODE:
        # 测试模式，只检查特定商品
        test_products = [
            # 商品货号, 单价
            ["703.786.83", 1599],  # 假设单价高于现价
            ["004.701.14", 399],   # 假设单价低于现价
            ["705.316.56", 2499],  # a
            ["102.635.24", 99]     # b
        ]
        
        print("\n=== 测试模式：开始测试商品价格获取 ===\n")
        results = []
        for product_info in test_products:
            product_number = product_info[0]
            unit_price = product_info[1]
            details = test_with_unit_price(product_number, unit_price)
            results.append((details, unit_price))
            time.sleep(2)
        
        print("\n=== 测试结果汇总 ===")
        for result, unit_price in results:
            current_price = result.get("current_price")
            if current_price:
                price_lower = "是" if current_price < unit_price else "否"
                print(f"商品 {result['product_number']}: " 
                      f"单价 ¥{unit_price:.2f}, "
                      f"现价 ¥{current_price:.2f}, "
                      f"现价低于单价: {price_lower}")
            else:
                print(f"商品 {result['product_number']}: 无法获取价格信息")
    else:
        # 正常模式，更新Excel
        print(f"开始更新Excel文件: {excel_file}")
        result = update_excel_prices(excel_file)
        
        if result:
            print("Excel更新成功！")
        else:
            print("Excel更新失败，请检查日志")